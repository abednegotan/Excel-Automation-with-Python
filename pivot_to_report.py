from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

month = input('Please enter the month :')

wb = load_workbook('pivot_table.xlsx')

sheet = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column

min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_col, max_col=max_col, min_row=min_row+1, max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12")

barchart.title = "Sales of different modes of shipping method"
barchart.style = 5

for i in range(min_col+1, max_col+1):
    
    letter = get_column_letter(i)
    
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'
    
    

sheet['A1'] = 'Sales Report'
sheet['A2'] = month

sheet['A1'].font = Font('Arial', bold = True, size = 20)
sheet['A2'].font = Font('Arial', bold = True, size = 10)

wb.save(f'report_{month}.xlsx')